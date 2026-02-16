from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .severity import SeverityRule
from .state_schema import SeverityLevel


@dataclass(frozen=True, slots=True)
class DomainQuestion:
    domain_id: str
    presenting_concern: str
    question: str
    trigger: str
    tag: str
    dsm_red_flag: str


def _norm(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _header_index(headers: list[str], candidates: tuple[str, ...]) -> int:
    normalized = [_norm(h) for h in headers]
    for i, h in enumerate(normalized):
        for c in candidates:
            if c in h:
                return i
    raise ValueError(f"Could not find header candidates {candidates}. Found: {headers}")


def _normalize_domain_id(sheet_name: str) -> str:
    return (
        sheet_name.strip()
        .lower()
        .replace("&", "and")
        .replace(",", "")
        .replace(" ", "_")
        .replace("__", "_")
    )


def load_domain_questions(workbook_path: Path | str) -> list[DomainQuestion]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    records: list[DomainQuestion] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(min_row=1, values_only=True)
        header_row = next(rows)
        headers = ["" if h is None else str(h).strip() for h in header_row]

        try:
            concern_i = _header_index(headers, ("presenting concern",))
            question_i = _header_index(headers, ("question", "question to ask", "question (llm asks)"))
            trigger_i = _header_index(headers, ("trigger", "response trigger"))
            tag_i = _header_index(headers, ("tag assigned", "ai tag", "primary tag"))
            red_flag_i = _header_index(headers, ("dsm", "risk indicator", "risk"))
        except ValueError:
            # skip sheets that are not domain question tables
            continue

        domain_id = _normalize_domain_id(sheet_name)
        for row in rows:
            values = ["" if v is None else str(v).strip() for v in row]
            question = values[question_i] if question_i < len(values) else ""
            if not question:
                continue

            record = DomainQuestion(
                domain_id=domain_id,
                presenting_concern=values[concern_i] if concern_i < len(values) else "",
                question=question,
                trigger=values[trigger_i] if trigger_i < len(values) else "",
                tag=values[tag_i] if tag_i < len(values) else "",
                dsm_red_flag=values[red_flag_i] if red_flag_i < len(values) else "",
            )
            records.append(record)

    if not records:
        raise ValueError("No domain question records were parsed from workbook")
    return records


def _parse_tier(value: str) -> SeverityLevel | None:
    v = _norm(value)
    if "tier 3" in v:
        return SeverityLevel.HIGH
    if "tier 2" in v:
        return SeverityLevel.MODERATE
    if "tier 1" in v:
        return SeverityLevel.MILD_CONCERN
    if "tier 0" in v:
        return SeverityLevel.LOW
    return None


def load_severity_rules(workbook_path: Path | str) -> list[SeverityRule]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if "Severity Escalation" not in wb.sheetnames:
        raise ValueError("Missing 'Severity Escalation' sheet")

    ws = wb["Severity Escalation"]
    rows = ws.iter_rows(min_row=1, values_only=True)
    headers = ["" if h is None else str(h).strip() for h in next(rows)]

    rule_id_i = _header_index(headers, ("rule id",))
    trigger_i = _header_index(headers, ("escalation trigger", "condition"))
    escalated_tier_i = _header_index(headers, ("escalated tier",))

    rules: list[SeverityRule] = []
    for row in rows:
        values = ["" if v is None else str(v).strip() for v in row]
        rule_id = values[rule_id_i] if rule_id_i < len(values) else ""
        trigger = values[trigger_i] if trigger_i < len(values) else ""
        outcome = _parse_tier(values[escalated_tier_i] if escalated_tier_i < len(values) else "")

        if not rule_id or not trigger or outcome is None:
            continue

        tags = [t.strip() for t in trigger.replace(";", ",").split(",") if t.strip()]
        rules.append(
            SeverityRule(
                rule_id=rule_id,
                trigger_tags=frozenset(tags),
                outcome=outcome,
                priority=1,
            )
        )

    if not rules:
        raise ValueError("No severity rules parsed from Severity Escalation sheet")
    return rules
