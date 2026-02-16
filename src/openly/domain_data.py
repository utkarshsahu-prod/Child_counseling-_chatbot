"""Comprehensive domain tree and cross-domain data models.

Parses all 9 domain sheets from Domain_tree_UPDATED.xlsx and all 7 sheets
from Cross_Domain_Logic_UPDATED.xlsx into typed runtime structures.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Domain Tree models
# ---------------------------------------------------------------------------

@dataclass(frozen=True, slots=True)
class ResponseTrigger:
    """A single response-trigger leaf in the domain tree."""
    trigger_text: str
    ai_tag: str
    dsm_red_flag: str


@dataclass(slots=True)
class DomainQuestion:
    """A question node with its child response triggers."""
    question_text: str
    triggers: list[ResponseTrigger] = field(default_factory=list)


@dataclass(slots=True)
class PresentingConcern:
    """A presenting concern grouping with its child questions."""
    name: str
    questions: list[DomainQuestion] = field(default_factory=list)

    @property
    def all_tags(self) -> list[str]:
        return [t.ai_tag for q in self.questions for t in q.triggers if t.ai_tag]

    @property
    def all_red_flags(self) -> list[str]:
        flags = []
        for q in self.questions:
            for t in q.triggers:
                if t.dsm_red_flag and t.dsm_red_flag != "—":
                    flags.append(t.dsm_red_flag)
        return flags


@dataclass(slots=True)
class ClinicalDomain:
    """A full clinical domain parsed from one Excel sheet."""
    domain_id: str
    display_name: str
    concerns: list[PresentingConcern] = field(default_factory=list)

    @property
    def all_tags(self) -> set[str]:
        return {tag for c in self.concerns for tag in c.all_tags}

    @property
    def all_questions(self) -> list[str]:
        return [q.question_text for c in self.concerns for q in c.questions]

    def find_concern_by_tag(self, tag: str) -> PresentingConcern | None:
        for c in self.concerns:
            if tag in c.all_tags:
                return c
        return None


# ---------------------------------------------------------------------------
# Cross-Domain Logic models
# ---------------------------------------------------------------------------

@dataclass(frozen=True, slots=True)
class ConvergencePattern:
    rule_id: str
    pattern_name: str
    clinical_hypothesis: str
    domain_tags: dict[str, list[str]]   # domain_source -> tags
    min_domains_required: int
    confidence_tier: str
    escalation_action: str
    recommended_evaluation: str
    notes: str


@dataclass(frozen=True, slots=True)
class ConfoundRule:
    rule_id: str
    confound_name: str
    apparent_hypothesis: str
    confounding_tags: list[str]
    confound_source: str
    confound_logic: str
    action: str
    parent_message: str
    clinical_rationale: str


@dataclass(frozen=True, slots=True)
class SeverityEscalation:
    rule_id: str
    trigger: str
    condition: str
    current_tier: str
    escalated_tier: str
    action: str
    urgency: str
    notes: str


@dataclass(frozen=True, slots=True)
class DifferentialRule:
    condition_a: str
    condition_b: str
    overlapping_tags: list[str]
    key_differentiator: str
    tags_favoring_a: list[str]
    tags_favoring_b: list[str]
    decision_rule: str


@dataclass(frozen=True, slots=True)
class AgeLogicRule:
    age_band: str
    domain: str
    tag_pattern: str
    age_normative: str
    age_concerning: str
    clinical_notes: str
    convergence_impact: str


@dataclass(frozen=True, slots=True)
class TagReference:
    ai_tag: str
    source_domain: str
    role: str
    connected_patterns: list[str]
    clinical_category: str
    notes: str


@dataclass(slots=True)
class CrossDomainData:
    """All cross-domain logic sheets parsed into one structure."""
    convergence_patterns: list[ConvergencePattern] = field(default_factory=list)
    confound_rules: list[ConfoundRule] = field(default_factory=list)
    severity_escalations: list[SeverityEscalation] = field(default_factory=list)
    differential_rules: list[DifferentialRule] = field(default_factory=list)
    age_logic_rules: list[AgeLogicRule] = field(default_factory=list)
    tag_reference: list[TagReference] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _s(val: object) -> str:
    """Safe string conversion."""
    if val is None:
        return ""
    return str(val).strip()


def _split_tags(text: str) -> list[str]:
    """Split newline/comma/semicolon-separated tag lists."""
    items = text.replace("\n", ",").replace(";", ",").split(",")
    return [t.strip() for t in items if t.strip()]


def _normalize_domain_id(sheet_name: str) -> str:
    return (
        sheet_name.strip()
        .lower()
        .replace("&", "and")
        .replace(",", "")
        .replace(" ", "_")
        .replace("__", "_")
        .rstrip("_")
    )


def _parse_min_domains(text: str) -> int:
    """Extract leading integer from strings like '3 of 4' or '2'."""
    import re
    m = re.match(r"(\d+)", text)
    return int(m.group(1)) if m else 2


def _find_col(headers: list[str], *candidates: str) -> int:
    """Find column index matching any candidate substring (case-insensitive)."""
    normed = [h.lower() for h in headers]
    for c in candidates:
        for i, h in enumerate(normed):
            if c in h:
                return i
    return -1


def _cell(row: tuple, idx: int) -> str:
    if idx < 0 or idx >= len(row):
        return ""
    return _s(row[idx])


# ---------------------------------------------------------------------------
# Domain tree parser
# ---------------------------------------------------------------------------

def load_all_domains(workbook_path: Path | str) -> list[ClinicalDomain]:
    """Parse all domain sheets from Domain_tree_UPDATED.xlsx."""
    wb = load_workbook(str(workbook_path), read_only=True, data_only=True)
    domains: list[ClinicalDomain] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(min_row=1, values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            continue

        headers = [_s(h) for h in header_row]

        # Find columns - headers vary across sheets
        concern_i = _find_col(headers, "presenting concern", "presenting concern (parent")
        question_i = _find_col(headers, "question to ask", "question (llm", "question")
        trigger_i = _find_col(headers, "parent response trigger", "parent response triggers", "response trigger")
        tag_i = _find_col(headers, "ai tag assigned", "primary tag assigned", "primary ai tag", "tag assigned")
        flag_i = _find_col(headers, "dsm-5 red flag", "dsm-5 risk", "dsm-5 indicator", "risk indicator")

        # Motor sheet has no "Presenting Concern" header - first real col is question
        if concern_i < 0 and question_i < 0:
            # Try treating column 0 as concern if header is blank/space
            if headers[0].strip() == "" or "presenting" in headers[0].lower():
                concern_i = 0
            # Column layout: sometimes concern is implicit
            if question_i < 0:
                question_i = _find_col(headers, "question")
        if trigger_i < 0:
            trigger_i = _find_col(headers, "trigger")
        if tag_i < 0:
            tag_i = _find_col(headers, "tag")
        if flag_i < 0:
            flag_i = _find_col(headers, "risk")

        domain = ClinicalDomain(
            domain_id=_normalize_domain_id(sheet_name),
            display_name=sheet_name.strip(),
        )

        current_concern: PresentingConcern | None = None
        current_question: DomainQuestion | None = None

        for row in rows_iter:
            vals = tuple(_s(c) for c in row)
            # Skip fully empty rows
            if not any(v for v in vals):
                continue

            concern_text = _cell(row, concern_i) if concern_i >= 0 else ""
            question_text = _cell(row, question_i)
            trigger_text = _cell(row, trigger_i)
            tag_text = _cell(row, tag_i)
            flag_text = _cell(row, flag_i)

            # New presenting concern
            if concern_text:
                current_concern = PresentingConcern(name=concern_text)
                domain.concerns.append(current_concern)
                current_question = None

            # Ensure we have a concern container
            if current_concern is None:
                current_concern = PresentingConcern(name="(uncategorized)")
                domain.concerns.append(current_concern)

            # New question
            if question_text:
                current_question = DomainQuestion(question_text=question_text)
                current_concern.questions.append(current_question)

            # Add trigger to current question
            if current_question is not None and trigger_text:
                current_question.triggers.append(
                    ResponseTrigger(
                        trigger_text=trigger_text,
                        ai_tag=tag_text,
                        dsm_red_flag=flag_text,
                    )
                )

        if domain.concerns:
            domains.append(domain)

    wb.close()
    return domains


# ---------------------------------------------------------------------------
# Cross-domain logic parser
# ---------------------------------------------------------------------------

def load_cross_domain_data(workbook_path: Path | str) -> CrossDomainData:
    """Parse all sheets from Cross_Domain_Logic_UPDATED.xlsx."""
    wb = load_workbook(str(workbook_path), read_only=True, data_only=True)
    data = CrossDomainData()

    # --- Convergence Patterns ---
    if "Convergence Patterns" in wb.sheetnames:
        ws = wb["Convergence Patterns"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            rule_id = _s(row[0]) if len(row) > 0 else ""
            if not rule_id:
                continue
            domain_tags: dict[str, list[str]] = {}
            # Columns: 3=D1 tags, 4=D1 source, 5=D2 tags, 6=D2 source, etc.
            for tag_col, src_col in [(3, 4), (5, 6), (7, 8), (9, 10)]:
                tags = _split_tags(_cell(row, tag_col))
                source = _cell(row, src_col)
                if tags and source:
                    domain_tags[source] = tags
            data.convergence_patterns.append(ConvergencePattern(
                rule_id=rule_id,
                pattern_name=_cell(row, 1),
                clinical_hypothesis=_cell(row, 2),
                domain_tags=domain_tags,
                min_domains_required=_parse_min_domains(_cell(row, 11)),
                confidence_tier=_cell(row, 12),
                escalation_action=_cell(row, 13),
                recommended_evaluation=_cell(row, 14),
                notes=_cell(row, 15),
            ))

    # --- Confound Rules ---
    if "Confound Rules" in wb.sheetnames:
        ws = wb["Confound Rules"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rule_id = _s(row[0]) if len(row) > 0 else ""
            if not rule_id:
                continue
            data.confound_rules.append(ConfoundRule(
                rule_id=rule_id,
                confound_name=_cell(row, 1),
                apparent_hypothesis=_cell(row, 2),
                confounding_tags=_split_tags(_cell(row, 3)),
                confound_source=_cell(row, 4),
                confound_logic=_cell(row, 5),
                action=_cell(row, 6),
                parent_message=_cell(row, 7),
                clinical_rationale=_cell(row, 8),
            ))

    # --- Severity Escalation ---
    if "Severity Escalation" in wb.sheetnames:
        ws = wb["Severity Escalation"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rule_id = _s(row[0]) if len(row) > 0 else ""
            if not rule_id:
                continue
            data.severity_escalations.append(SeverityEscalation(
                rule_id=rule_id,
                trigger=_cell(row, 1),
                condition=_cell(row, 2),
                current_tier=_cell(row, 3),
                escalated_tier=_cell(row, 4),
                action=_cell(row, 5),
                urgency=_cell(row, 6),
                notes=_cell(row, 7),
            ))

    # --- Differential Diagnosis ---
    if "Differential Diagnosis" in wb.sheetnames:
        ws = wb["Differential Diagnosis"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            cond_a = _s(row[0]) if len(row) > 0 else ""
            if not cond_a:
                continue
            data.differential_rules.append(DifferentialRule(
                condition_a=cond_a,
                condition_b=_cell(row, 1),
                overlapping_tags=_split_tags(_cell(row, 2)),
                key_differentiator=_cell(row, 3),
                tags_favoring_a=_split_tags(_cell(row, 4)),
                tags_favoring_b=_split_tags(_cell(row, 5)),
                decision_rule=_cell(row, 6),
            ))

    # --- Age Logic Layer ---
    if "Age Logic Layer" in wb.sheetnames:
        ws = wb["Age Logic Layer"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            age_band = _s(row[0]) if len(row) > 0 else ""
            if not age_band:
                continue
            data.age_logic_rules.append(AgeLogicRule(
                age_band=age_band,
                domain=_cell(row, 1),
                tag_pattern=_cell(row, 2),
                age_normative=_cell(row, 3),
                age_concerning=_cell(row, 4),
                clinical_notes=_cell(row, 5),
                convergence_impact=_cell(row, 6),
            ))

    # --- Tag Reference Map ---
    if "Tag Reference Map" in wb.sheetnames:
        ws = wb["Tag Reference Map"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            ai_tag = _s(row[0]) if len(row) > 0 else ""
            if not ai_tag:
                continue
            data.tag_reference.append(TagReference(
                ai_tag=ai_tag,
                source_domain=_cell(row, 1),
                role=_cell(row, 2),
                connected_patterns=_split_tags(_cell(row, 3)),
                clinical_category=_cell(row, 4),
                notes=_cell(row, 5),
            ))

    wb.close()
    return data
