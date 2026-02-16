from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


@dataclass(frozen=True, slots=True)
class SheetContract:
    sheet_name: str
    required_columns: tuple[str, ...]
    required_non_empty_columns: tuple[str, ...] = ()


@dataclass(frozen=True, slots=True)
class IngestionReport:
    sheet_name: str
    records_parsed: int


class ContractError(ValueError):
    pass


def _normalize(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _sheet_header_index(sheet_headers: Iterable[object]) -> dict[str, int]:
    seen: dict[str, int] = {}
    for index, header in enumerate(sheet_headers):
        key = _normalize(header)
        if not key:
            continue
        if key in seen:
            raise ContractError(f"Duplicate header detected: {header}")
        seen[key] = index
    return seen


def validate_sheet_headers(workbook_path: Path | str, contract: SheetContract) -> dict[str, int]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if contract.sheet_name not in wb.sheetnames:
        raise ContractError(f"Missing sheet: {contract.sheet_name}")

    ws = wb[contract.sheet_name]
    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = _sheet_header_index(header_cells)

    missing = [col for col in contract.required_columns if _normalize(col) not in header_map]
    if missing:
        raise ContractError(
            f"Sheet '{contract.sheet_name}' is missing required columns: {', '.join(missing)}"
        )
    return header_map


def convert_workbook_to_records(
    workbook_path: Path | str,
    sheet_name: str,
    required_columns: Iterable[str],
    required_non_empty_columns: Iterable[str] = (),
) -> list[dict]:
    contract = SheetContract(
        sheet_name=sheet_name,
        required_columns=tuple(required_columns),
        required_non_empty_columns=tuple(required_non_empty_columns),
    )
    validate_sheet_headers(workbook_path, contract)

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    rows = ws.iter_rows(min_row=1, values_only=True)
    headers = next(rows)
    normalized_to_index = _sheet_header_index(headers)

    out: list[dict] = []
    for row_num, row in enumerate(rows, start=2):
        record = {}
        is_empty = True
        for col in contract.required_columns:
            value = row[normalized_to_index[_normalize(col)]]
            if value not in (None, ""):
                is_empty = False
            record[col] = value

        if is_empty:
            continue

        for col in contract.required_non_empty_columns:
            value = record.get(col)
            if value in (None, ""):
                raise ContractError(
                    f"Sheet '{sheet_name}' row {row_num} has empty required value for column '{col}'"
                )

        out.append(record)

    if not out:
        raise ContractError(f"Sheet '{sheet_name}' contains no valid records")
    return out


def ingest_sheet_with_contract(workbook_path: Path | str, contract: SheetContract) -> tuple[list[dict], IngestionReport]:
    records = convert_workbook_to_records(
        workbook_path,
        sheet_name=contract.sheet_name,
        required_columns=contract.required_columns,
        required_non_empty_columns=contract.required_non_empty_columns,
    )
    return records, IngestionReport(sheet_name=contract.sheet_name, records_parsed=len(records))
